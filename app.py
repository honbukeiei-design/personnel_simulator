from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    import plotly.express as px
    import plotly.graph_objects as go
except Exception:  # pragma: no cover
    px = None
    go = None

APP_TITLE = "病床機能・経営データ統合 人員要望意思決定支援"
DEFAULT_FINANCIAL_FILE = Path("data/financial_hospital_panel.xlsx")
DEFAULT_APPLICATION_FILE = Path("data/application_form.xlsx")
MHLW_BED_FUNCTION_TOP = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000055891.html"
MHLW_R6_OPEN_DATA = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/open_data_00018.html"
MHLW_R6_CHUBU_WARD_STYLE1 = "https://www.mhlw.go.jp/content/10800000/001299895.xlsx"
MHLW_R6_CHUBU_WARD_STYLE2_ANNUAL = "https://www.mhlw.go.jp/content/10800000/001299937.xlsx"
JILPT_LABOR_FORCE_URL = "https://www.jil.go.jp/kokunai/statistics/timeseries/html/g0201.html"
STAT_CPI_URL = "https://www.stat.go.jp/data/cpi/"
STAT_CPI_LATEST_URL = "https://www.stat.go.jp/data/cpi/sokuhou/tsuki/index-z.html"

JOB_MAP = {
    "医師": ["医師", "doctor", "physician"],
    "看護師": ["看護師", "看護", "nurse"],
    "医療スタッフ": ["医療スタッフ", "コメディカル", "薬剤", "検査", "放射線", "リハ", "技師", "med"],
    "事務職員": ["事務", "clerical", "admin"],
    "その他職員": ["その他"],
}
STAFF_COL_BY_JOB = {
    "医師": "医師数",
    "看護師": "看護師数",
    "医療スタッフ": "医療スタッフ数",
    "事務職員": "事務職員数",
    "その他職員": "その他職員数",
}
DEFAULT_ANNUAL_COST_YEN = {
    "医師": 18_000_000,
    "看護師": 6_500_000,
    "医療スタッフ": 6_000_000,
    "事務職員": 4_500_000,
    "その他職員": 4_500_000,
}


def normalize_text(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", "", text)


def yen_man_to_yen(value: float | int | None) -> float:
    """添付経営データは多くの金額列が千円単位で入っている想定。"""
    if pd.isna(value):
        return 0.0
    return float(value) * 1000


def safe_div(numer: pd.Series | float, denom: pd.Series | float) -> pd.Series | float:
    return np.where(np.asarray(denom, dtype="float64") == 0, np.nan, np.asarray(numer, dtype="float64") / np.asarray(denom, dtype="float64"))


@st.cache_data(show_spinner=False)
def load_financial_excel(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    numeric_candidates = [
        "年度", "修正医業収益", "医業費用", "人件費", "医薬品費", "医療材料費", "医業利益", "経常利益", "病床数", "全職員数",
        "事務職員数", "医師数", "看護師数", "医療スタッフ数", "その他職員数", "一日平均入院患者数", "一日平均外来患者数",
        "延べ入院患者数", "延べ外来患者数", "DPCフラグ", "地域医療支援病院フラグ", "Ta", "Da", "profit_deficit", "クラスタ",
    ]
    for col in numeric_candidates:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "病院名称" in df.columns:
        df["_hospital_key"] = df["病院名称"].map(normalize_text)
    if "年度" in df.columns:
        df["年度"] = df["年度"].fillna(0).astype(int)
    return df


@st.cache_data(show_spinner=False)
def load_application_excel(file_bytes: bytes) -> Dict[str, object]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    def v(cell: str) -> str:
        val = ws[cell].value
        return "" if val is None else str(val).strip()

    def collect_range(range_addr: str) -> str:
        vals: List[str] = []
        for row in ws[range_addr]:
            for cell in row:
                if cell.value not in (None, ""):
                    vals.append(str(cell.value).strip())
        return "\n".join(dict.fromkeys(vals))

    # 申請表は結合セルが多いため、ラベル周辺の入力候補セルを広めに拾う。
    raw_job = collect_range("J11:Q12")
    raw_count = collect_range("W11:AD12") or collect_range("R11:V12")
    raw_period = collect_range("W15:AD17") or collect_range("R15:V17")
    reason = collect_range("J23:AD26") or collect_range("J20:AD26")
    constraint = collect_range("J29:AD32")
    comment_nursing = collect_range("J50:AD53")
    comment_admin = collect_range("J56:AD59")
    comment_director = collect_range("J62:AD65")
    comment_hq = collect_range("J68:AD71")

    # チェックボックスや丸印が入力されることを想定し、候補ラベルの周辺を保持。
    impact_area = {
        "増収効果あり": collect_range("K35:O44"),
        "現状維持": collect_range("Q35:U44"),
        "コスト増だが必要不可欠": collect_range("W35:AD44"),
    }
    category_area = {
        "新規増員": collect_range("K20:Q21"),
        "欠員補充": collect_range("S20:V21"),
    }
    hire_type_area = {
        "機構採用": collect_range("J15:Q17"),
        "病院採用": collect_range("A15:I17"),
    }

    count_num = None
    m = re.search(r"\d+(?:\.\d+)?", raw_count)
    if m:
        count_num = float(m.group())

    return {
        "申請日": collect_range("X4:AD4") or v("X4"),
        "所属名": collect_range("X6:AD6") or v("X6"),
        "氏名": collect_range("X7:AD7") or v("X7"),
        "希望職種_raw": raw_job,
        "希望人数_raw": raw_count,
        "希望人数": count_num,
        "採用時期": raw_period,
        "採用区分_候補": hire_type_area,
        "申請区分_候補": category_area,
        "希望理由": reason,
        "既存人員で対応できない理由": constraint,
        "収支影響_候補": impact_area,
        "看護部長等コメント": comment_nursing,
        "事務部長コメント": comment_admin,
        "院長コメント": comment_director,
        "本部コメント": comment_hq,
    }


def classify_job(raw: str) -> str:
    key = normalize_text(raw).lower()
    for job, words in JOB_MAP.items():
        if any(normalize_text(w).lower() in key for w in words):
            return job
    return "看護師"


def read_file_bytes(uploaded, default_path: Path) -> Optional[bytes]:
    """Return uploaded bytes, bundled default bytes, or None when neither exists.

    Streamlit Cloud deployments should normally *not* include private xlsx files.
    In that case, the app must wait for user upload instead of trying to read
    a missing local file and raising FileNotFoundError.
    """
    if uploaded is not None:
        return uploaded.getvalue()
    if default_path.exists():
        return default_path.read_bytes()
    return None


def require_excel_bytes(file_bytes: Optional[bytes], label: str) -> bytes:
    if file_bytes is None:
        st.info(f"{label}をアップロードしてください。GitHub/Streamlit Cloudには個人情報・機微データを置かない設計です。")
        st.stop()
    return file_bytes


def latest_rows(df: pd.DataFrame) -> pd.DataFrame:
    if "年度" not in df.columns or "病院名称" not in df.columns:
        return df.copy()
    return df.sort_values("年度").groupby("_hospital_key", as_index=False).tail(1)


def add_core_metrics(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    # 添付データの金額列は千円単位を想定し、円換算列を追加する。
    for col in ["修正医業収益", "医業費用", "人件費", "医薬品費", "医療材料費", "医業利益", "経常利益"]:
        if col in out.columns:
            out[col + "_円"] = out[col].map(yen_man_to_yen)

    if {"修正医業収益_円", "医薬品費_円", "医療材料費_円"}.issubset(out.columns):
        out["粗利益_円"] = out["修正医業収益_円"] - out["医薬品費_円"] - out["医療材料費_円"]
    if {"粗利益_円", "人件費_円"}.issubset(out.columns):
        out["粗利益人件費カバー率"] = safe_div(out["粗利益_円"], out["人件費_円"])
        out["粗利益人件費余力_円"] = out["粗利益_円"] - out["人件費_円"]
        out["人件費対粗利益率"] = safe_div(out["人件費_円"], out["粗利益_円"])

    if {"一日平均入院患者数", "病床数"}.issubset(out.columns):
        out["病床稼働率"] = safe_div(out["一日平均入院患者数"], out["病床数"])
    if {"全職員数", "病床数"}.issubset(out.columns):
        out["職員数_対病床"] = safe_div(out["全職員数"], out["病床数"])
    for job, col in STAFF_COL_BY_JOB.items():
        if {col, "病床数"}.issubset(out.columns):
            out[f"{job}_対病床"] = safe_div(out[col], out["病床数"])
    if {"修正医業収益_円", "延べ入院患者数"}.issubset(out.columns):
        out["入院患者1人日あたり医業収益_円"] = safe_div(out["修正医業収益_円"], out["延べ入院患者数"])
    if {"人件費_円", "全職員数"}.issubset(out.columns):
        out["職員1人あたり人件費_円"] = safe_div(out["人件費_円"], out["全職員数"])
    if {"医業利益_円", "修正医業収益_円"}.issubset(out.columns):
        out["医業利益率_calc"] = safe_div(out["医業利益_円"], out["修正医業収益_円"])
    return out

def fuzzy_pick_hospital(df: pd.DataFrame, applicant_dept: str = "", default_keyword: str = "長野県立") -> str:
    names = sorted(df["病院名称"].dropna().astype(str).unique()) if "病院名称" in df.columns else []
    if not names:
        return ""
    dept_key = normalize_text(applicant_dept)
    for n in names:
        if normalize_text(n) and normalize_text(n) in dept_key:
            return n
    for n in names:
        if default_keyword in str(n):
            return n
    return names[0]


def peer_group(
    df_latest: pd.DataFrame,
    target_row: pd.Series,
    criteria: List[str] | None = None,
    min_peers: int = 5,
) -> Tuple[pd.DataFrame, List[str]]:
    """選択された属性でピア群を絞り込む。

    絞り込み過ぎて対象が少なくなると中央値が不安定になるため、各条件は
    適用後に min_peers 以上残る場合だけ採用する。
    """
    criteria = criteria or ["クラスタ", "DPCフラグ", "過疎地区分", "小児フラグ", "精神科フラグ"]
    peers = df_latest.copy()
    applied: List[str] = []
    for col in criteria:
        if col in peers.columns and col in target_row.index and pd.notna(target_row.get(col)):
            subset = peers[peers[col] == target_row.get(col)]
            if len(subset) >= min_peers:
                peers = subset
                applied.append(f"{col}={target_row.get(col)}")
    if peers.empty:
        return df_latest.copy(), []
    return peers, applied


def selected_key(option_dict: Dict[str, object]) -> str:
    """申請フォームのラジオ選択値を安全に取り出す。"""
    if isinstance(option_dict, dict) and option_dict:
        return next(iter(option_dict.keys()))
    return ""


def application_adjustments(hire_type: str, request_type: str, revenue_impact: str) -> Dict[str, float | str]:
    """申請フォームの区分をシミュレーション係数へ反映する。

    反映方針:
    - 増収効果あり: 年間増収推計と財務スコアを加点
    - 現状維持: 収益よりも医療提供体制維持・質安全の価値を厚く評価
    - コスト増だが必要不可欠: 政策・質安全価値と人員不足解消を厚く評価しつつ、財務面は慎重に扱う
    - 欠員補充: 新規の収益創出よりもサービス維持・採用難リスクを評価
    - 採用区分未定: 実行可能性リスクとして少し減点
    """
    revenue_multiplier = 1.0
    policy_weight_add = 0.0
    labor_factor_add = 0.0
    finance_score_bonus = 0.0
    policy_score_bonus = 0.0
    staff_gap_score_bonus = 0.0
    execution_penalty = 0.0
    notes: list[str] = []

    if revenue_impact == "増収効果あり":
        revenue_multiplier += 0.30
        finance_score_bonus += 1.2
        notes.append("収支影響=増収効果あり: 年間増収推計を30%上乗せし、財務スコアを加点")
    elif revenue_impact == "現状維持":
        policy_weight_add += 0.20
        policy_score_bonus += 1.5
        notes.append("収支影響=現状維持: 増収よりも医療提供体制維持として政策・質安全価値を加重")
    elif revenue_impact == "コスト増だが必要不可欠":
        policy_weight_add += 0.35
        policy_score_bonus += 2.0
        staff_gap_score_bonus += 0.8
        revenue_multiplier *= 0.70
        notes.append("収支影響=コスト増だが必要不可欠: 財務増収は控えめにし、必要不可欠性・人員不足解消を加点")

    if request_type == "新規増員":
        staff_gap_score_bonus += 0.4
        notes.append("申請区分=新規増員: ピア不足解消の評価を小幅加点")
    elif request_type == "欠員補充":
        policy_weight_add += 0.10
        labor_factor_add += 0.15
        notes.append("申請区分=欠員補充: 現状維持・欠員長期化リスクを加味")
    elif request_type == "その他":
        policy_weight_add += 0.05
        notes.append("申請区分=その他: 個別事情として政策価値を小幅加味")

    if hire_type == "機構採用":
        labor_factor_add += 0.05
        notes.append("採用区分=機構採用: 広域採用調整の必要性として採用市場リスクを小幅加味")
    elif hire_type == "病院採用":
        execution_penalty -= 0.2
        notes.append("採用区分=病院採用: 病院裁量で実行しやすい前提として小幅加点")
    elif hire_type == "未定":
        execution_penalty += 0.6
        notes.append("採用区分=未定: 実行可能性が未確定のため最終スコアを減点")

    return {
        "revenue_multiplier": revenue_multiplier,
        "policy_weight_add": policy_weight_add,
        "labor_factor_add": labor_factor_add,
        "finance_score_bonus": finance_score_bonus,
        "policy_score_bonus": policy_score_bonus,
        "staff_gap_score_bonus": staff_gap_score_bonus,
        "execution_penalty": execution_penalty,
        "notes": "\n".join(notes) if notes else "申請区分による補正なし",
    }

def simulate(
    target: pd.Series,
    peers: pd.DataFrame,
    job: str,
    add_count: float,
    annual_cost: float,
    revenue_sensitivity: float,
    quality_weight: float,
    labor_shortage_factor: float = 0.0,
    labor_weight: float = 0.15,
    inflation_rate: float = 0.0,
    inflation_cost_share: float = 0.6,
    hire_type: str = "",
    request_type: str = "",
    revenue_impact: str = "",
) -> Dict[str, float | str]:
    staff_col = STAFF_COL_BY_JOB.get(job, "看護師数")
    bed_count = float(target.get("病床数", np.nan) or np.nan)
    current_staff = float(target.get(staff_col, np.nan) or np.nan)
    current_ratio = current_staff / bed_count if bed_count else np.nan
    ratio_col = f"{job}_対病床"
    peer_median = float(peers[ratio_col].median()) if ratio_col in peers.columns and peers[ratio_col].notna().any() else np.nan
    gap_to_peer = (peer_median - current_ratio) * bed_count if pd.notna(peer_median) and pd.notna(current_ratio) else np.nan
    after_staff = current_staff + add_count if pd.notna(current_staff) else add_count
    after_ratio = after_staff / bed_count if bed_count else np.nan

    # 労働市場が逼迫しているほど、今採用しないことの機会損失を大きく見る。
    labor_factor = max(0.0, min(1.0, float(labor_shortage_factor)))
    inflation = max(-0.05, min(0.20, float(inflation_rate)))
    inflation_share = max(0.0, min(1.0, float(inflation_cost_share)))

    app_adj = application_adjustments(hire_type, request_type, revenue_impact)
    effective_quality_weight = max(0.0, min(1.5, quality_weight + float(app_adj["policy_weight_add"])))
    effective_labor_factor = max(0.0, min(1.0, labor_factor + float(app_adj["labor_factor_add"])))

    annual_cost_total_nominal = add_count * annual_cost
    annual_cost_total = annual_cost_total_nominal * (1 + inflation * inflation_share)
    revenue_base = float(target.get("修正医業収益_円", 0) or 0)
    med_income_per_day = float(target.get("入院患者1人日あたり医業収益_円", np.nan) or np.nan)
    if pd.notna(gap_to_peer) and gap_to_peer > 0 and add_count > 0:
        closing_ratio = min(add_count / gap_to_peer, 1.0)
    else:
        closing_ratio = 0.0
    annual_revenue_gain_base = revenue_base * revenue_sensitivity * closing_ratio
    annual_revenue_gain = annual_revenue_gain_base * float(app_adj["revenue_multiplier"])

    gross_profit = float(target.get("粗利益_円", np.nan) or np.nan)
    current_labor_cost = float(target.get("人件費_円", np.nan) or np.nan)
    after_labor_cost = current_labor_cost + annual_cost_total if pd.notna(current_labor_cost) else np.nan
    gross_profit_after_gain = gross_profit + annual_revenue_gain if pd.notna(gross_profit) else np.nan
    gross_profit_cover_after = gross_profit_after_gain / after_labor_cost if pd.notna(gross_profit_after_gain) and after_labor_cost else np.nan
    gross_profit_margin_after = gross_profit_after_gain - after_labor_cost if pd.notna(gross_profit_after_gain) and pd.notna(after_labor_cost) else np.nan

    policy_value = annual_cost_total * effective_quality_weight
    # 労働市場逼迫価値：将来採用難化・紹介料増・欠員長期化のリスクをコスト換算で見える化。
    labor_market_value = annual_cost_total * labor_weight * effective_labor_factor
    net_financial = annual_revenue_gain - annual_cost_total
    net_with_policy = annual_revenue_gain + policy_value + labor_market_value - annual_cost_total
    roi_months = np.nan if annual_revenue_gain <= 0 else annual_cost_total / annual_revenue_gain * 12

    score_staff_gap = 0 if pd.isna(gap_to_peer) else max(0, min(10, gap_to_peer / max(add_count, 1) * 5 + float(app_adj["staff_gap_score_bonus"])))
    score_finance = max(0, min(10, (net_with_policy / max(annual_cost_total, 1) + 1) * 5 + float(app_adj["finance_score_bonus"]))) if annual_cost_total else 0
    if pd.isna(gross_profit_cover_after):
        score_sustainability = 5
        sustainability_label = "粗利益・人件費データ不足のため要確認"
    elif gross_profit_cover_after >= 1.15 and gross_profit_margin_after >= 0:
        score_sustainability = 9
        sustainability_label = "継続可能：粗利益で増員後人件費を概ね支えられる"
    elif gross_profit_cover_after >= 1.0 and gross_profit_margin_after >= 0:
        score_sustainability = 7
        sustainability_label = "注意付き継続可能：余力は小さい"
    elif gross_profit_cover_after >= 0.9:
        score_sustainability = 4
        sustainability_label = "要条件整理：粗利益で人件費を支える余力が不足気味"
    else:
        score_sustainability = 2
        sustainability_label = "慎重判断：粗利益による人件費支持が弱い"
    score_policy = 8 if effective_quality_weight >= 0.5 else (6 if effective_quality_weight >= 0.25 else 4)
    score_policy = max(0, min(10, score_policy + float(app_adj["policy_score_bonus"])))
    score_labor = effective_labor_factor * 10
    raw_decision_score = (
        0.30 * score_staff_gap
        + 0.25 * score_finance
        + 0.25 * score_sustainability
        + 0.10 * score_policy
        + 0.10 * score_labor
        - float(app_adj["execution_penalty"])
    )
    decision_score = round(max(0, min(10, raw_decision_score)), 1)
    if decision_score >= 7.5:
        recommendation = "承認候補：人員不足・粗利益持続性・政策/採用市場リスクを説明しやすい"
    elif decision_score >= 5.5:
        recommendation = "条件付き承認候補：採用時期・人数・粗利益改善策・効果指標を明確化"
    else:
        recommendation = "再検討候補：効果仮説、配置計画、または人件費支持力の追加説明が必要"

    return {
        "現員": current_staff,
        "増員後": after_staff,
        "現員_対病床": current_ratio,
        "増員後_対病床": after_ratio,
        "ピア中央値_対病床": peer_median,
        "ピア中央値までの不足人数": gap_to_peer,
        "年間人件費増_物価反映前": annual_cost_total_nominal,
        "年間人件費増": annual_cost_total,
        "年間増収推計": annual_revenue_gain,
        "年間増収推計_申請補正前": annual_revenue_gain_base,
        "申請補正_増収倍率": float(app_adj["revenue_multiplier"]),
        "政策・質安全価値_実効重み": effective_quality_weight,
        "労働市場逼迫度_実効値": effective_labor_factor,
        "申請区分補正メモ": str(app_adj["notes"]),
        "現在粗利益": gross_profit,
        "現在人件費": current_labor_cost,
        "増員後粗利益": gross_profit_after_gain,
        "増員後人件費": after_labor_cost,
        "増員後粗利益人件費カバー率": gross_profit_cover_after,
        "増員後粗利益人件費余力": gross_profit_margin_after,
        "粗利益持続性判定": sustainability_label,
        "政策・質安全価値": policy_value,
        "労働市場逼迫価値": labor_market_value,
        "財務純効果": net_financial,
        "政策・労働市場価値込み純効果": net_with_policy,
        "ROI回収月数": roi_months,
        "意思決定スコア": decision_score,
        "推奨判定": recommendation,
        "患者1人日単価": med_income_per_day,
    }


@st.cache_data(show_spinner=False)
def scrape_mhlw_links() -> pd.DataFrame:
    try:
        html = requests.get(MHLW_R6_OPEN_DATA, timeout=20).text
        soup = BeautifulSoup(html, "html.parser")
        rows = []
        for a in soup.find_all("a"):
            href = a.get("href", "")
            text = a.get_text(" ", strip=True)
            if ".xlsx" in href.lower():
                if href.startswith("/"):
                    href = "https://www.mhlw.go.jp" + href
                rows.append({"表示名": text or href.rsplit("/", 1)[-1], "URL": href})
        return pd.DataFrame(rows).drop_duplicates()
    except Exception as exc:
        return pd.DataFrame([{"表示名": "取得失敗", "URL": str(exc)}])


@st.cache_data(show_spinner=False)
def try_load_mhlw_excel(url: str, max_rows: int = 5000) -> pd.DataFrame:
    try:
        res = requests.get(url, timeout=60)
        res.raise_for_status()
        return pd.read_excel(io.BytesIO(res.content), nrows=max_rows)
    except Exception as exc:
        return pd.DataFrame({"取得エラー": [str(exc)], "URL": [url]})


def metric_card(label: str, value: str, help_text: str | None = None):
    st.metric(label, value, help=help_text)



def application_input_panel() -> Dict[str, object]:
    """01_（別紙1）申請表の主要項目をUIで入力・確認する。"""
    st.subheader("01_（別紙1）申請表 入力フォーム")
    st.caption("申請表Excelはアップロード不要です。会議で確認したい項目をこの画面で入力し、意思決定ダッシュボードへ反映します。")

    with st.form("application_form_ui"):
        c1, c2, c3 = st.columns(3)
        with c1:
            app_date = st.date_input("申請日")
            dept = st.text_input("所属名 / 申請病院名", placeholder="例：長野県立○○病院")
            applicant = st.text_input("申請者氏名")
        with c2:
            job = st.selectbox("希望職種", list(STAFF_COL_BY_JOB.keys()), index=1)
            count = st.number_input("希望人数", min_value=0.0, max_value=100.0, value=1.0, step=0.5)
            period = st.text_input("採用希望時期", placeholder="例：令和8年4月、可能な限り早期")
        with c3:
            hire_type = st.radio("採用区分", ["機構採用", "病院採用", "未定"], horizontal=True)
            request_type = st.radio("申請区分", ["新規増員", "欠員補充", "その他"], horizontal=True)
            revenue_impact = st.radio("収支影響", ["増収効果あり", "現状維持", "コスト増だが必要不可欠"], horizontal=False)

        reason = st.text_area("希望理由", height=110, placeholder="増員が必要な診療機能、患者数、施設基準、地域医療上の必要性など")
        constraint = st.text_area("既存人員で対応できない理由", height=90, placeholder="現員配置、夜勤・当直、業務量、兼務状況、採用難など")
        n1, n2 = st.columns(2)
        with n1:
            nursing_comment = st.text_area("看護部長等コメント", height=80)
            admin_comment = st.text_area("事務部長コメント", height=80)
        with n2:
            director_comment = st.text_area("院長コメント", height=80)
            hq_comment = st.text_area("本部コメント", height=80)
        submitted = st.form_submit_button("申請内容を反映")

    app = {
        "申請日": str(app_date),
        "所属名": dept,
        "氏名": applicant,
        "希望職種_raw": job,
        "希望人数_raw": f"{count:g}",
        "希望人数": float(count),
        "採用時期": period,
        "採用区分_候補": {hire_type: "選択"},
        "申請区分_候補": {request_type: "選択"},
        "希望理由": reason,
        "既存人員で対応できない理由": constraint,
        "収支影響_候補": {revenue_impact: "選択"},
        "看護部長等コメント": nursing_comment,
        "事務部長コメント": admin_comment,
        "院長コメント": director_comment,
        "本部コメント": hq_comment,
    }
    return app



def fmt_yen(value: object) -> str:
    try:
        if pd.isna(value):
            return "-"
        return f"{float(value):,.0f}円"
    except Exception:
        return "-"


def fmt_pct(value: object, digits: int = 1) -> str:
    try:
        if pd.isna(value):
            return "-"
        return f"{float(value) * 100:.{digits}f}%"
    except Exception:
        return "-"


def risk_flags(result: Dict[str, object], labor_shortage_factor: float, inflation_rate: float) -> List[Tuple[str, str]]:
    flags: List[Tuple[str, str]] = []
    cover = result.get("増員後粗利益人件費カバー率", np.nan)
    margin = result.get("増員後粗利益人件費余力", np.nan)
    if pd.notna(cover) and float(cover) < 1.0:
        flags.append(("🔴", "増員後、粗利益だけでは人件費を支え切れない可能性があります。"))
    elif pd.notna(cover) and float(cover) < 1.15:
        flags.append(("🟠", "粗利益による人件費カバー余力が小さく、採用人数・時期の条件整理が必要です。"))
    if pd.notna(margin) and float(margin) < 0:
        flags.append(("🔴", "粗利益人件費余力がマイナスです。増収効果か費用抑制の根拠確認が必要です。"))
    if labor_shortage_factor >= 0.7:
        flags.append(("🟠", "労働市場の逼迫度が高く、採用遅延・欠員長期化リスクがあります。"))
    if inflation_rate >= 0.03:
        flags.append(("🟠", "物価・賃金上昇圧力が高めで、翌年度以降の人件費上振れに注意が必要です。"))
    roi = result.get("ROI回収月数", np.nan)
    if pd.notna(roi) and float(roi) > 36:
        flags.append(("🟡", "ROI回収期間が3年超です。政策必要性・質安全価値の説明が重要です。"))
    if not flags:
        flags.append(("🟢", "重大な警告はありません。前提条件を確認しつつ承認判断に進めます。"))
    return flags


def make_score_breakdown(result: Dict[str, object]) -> pd.DataFrame:
    annual_cost = float(result.get("年間人件費増", 0) or 0)
    net = float(result.get("政策・労働市場価値込み純効果", 0) or 0)
    cover = result.get("増員後粗利益人件費カバー率", np.nan)
    gap = result.get("ピア中央値までの不足人数", np.nan)
    policy_weight = float(result.get("政策・質安全価値_実効重み", 0) or 0)
    labor_factor = float(result.get("労働市場逼迫度_実効値", 0) or 0)

    staff_score = 0 if pd.isna(gap) else max(0, min(10, float(gap) / max(1.0, float(result.get("増員後", 0) or 1)) * 5))
    finance_score = max(0, min(10, (net / max(annual_cost, 1) + 1) * 5)) if annual_cost else 0
    if pd.isna(cover):
        sustain_score = 5
    elif float(cover) >= 1.15:
        sustain_score = 9
    elif float(cover) >= 1.0:
        sustain_score = 7
    elif float(cover) >= 0.9:
        sustain_score = 4
    else:
        sustain_score = 2
    policy_score = 8 if policy_weight >= 0.5 else (6 if policy_weight >= 0.25 else 4)
    labor_score = labor_factor * 10
    return pd.DataFrame({
        "指標": ["人員不足解消", "財務・政策込み効果", "粗利益持続性", "政策・質安全", "採用市場リスク"],
        "スコア": [staff_score, finance_score, sustain_score, policy_score, labor_score],
    })


def render_visual_dashboard(result: Dict[str, object], job: str, hospital: str, comp: pd.DataFrame, labor_shortage_factor: float, inflation_rate: float) -> None:
    st.subheader("意思決定サマリ")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("採用推奨度", f"{result['意思決定スコア']}/10")
    k2.metric("粗利益/人件費", "-" if pd.isna(result['増員後粗利益人件費カバー率']) else f"{result['増員後粗利益人件費カバー率']:.2f}倍")
    k3.metric("粗利益余力", fmt_yen(result['増員後粗利益人件費余力']))
    k4.metric("年間純効果", fmt_yen(result['政策・労働市場価値込み純効果']))
    st.success(result["推奨判定"])

    flags = risk_flags(result, labor_shortage_factor, inflation_rate)
    with st.container(border=True):
        st.markdown("**危険信号・確認ポイント**")
        for icon, text in flags:
            st.write(f"{icon} {text}")

    if go is not None:
        gcol, mcol = st.columns([0.9, 1.1])
        with gcol:
            gauge = go.Figure(go.Indicator(
                mode="gauge+number",
                value=float(result["意思決定スコア"]),
                number={"suffix": " / 10"},
                gauge={"axis": {"range": [0, 10]}, "bar": {"thickness": 0.28}, "steps": [
                    {"range": [0, 5.5]}, {"range": [5.5, 7.5]}, {"range": [7.5, 10]}
                ]},
                title={"text": "採用推奨度"},
            ))
            gauge.update_layout(height=280, margin=dict(l=20, r=20, t=45, b=10))
            st.plotly_chart(gauge, use_container_width=True)
        with mcol:
            x = 0 if pd.isna(result['増員後粗利益人件費カバー率']) else max(0, min(10, float(result['増員後粗利益人件費カバー率']) * 5))
            y = max(0, min(10, float(result.get('政策・質安全価値_実効重み', 0) or 0) * 10 + 3))
            matrix = go.Figure()
            matrix.add_shape(type="rect", x0=0, x1=5, y0=0, y1=5, opacity=0.08)
            matrix.add_shape(type="rect", x0=5, x1=10, y0=0, y1=5, opacity=0.08)
            matrix.add_shape(type="rect", x0=0, x1=5, y0=5, y1=10, opacity=0.08)
            matrix.add_shape(type="rect", x0=5, x1=10, y0=5, y1=10, opacity=0.08)
            matrix.add_trace(go.Scatter(x=[x], y=[y], mode="markers+text", text=[hospital], textposition="top center", marker={"size": 18}))
            matrix.add_annotation(x=7.5, y=9.4, text="積極採用", showarrow=False)
            matrix.add_annotation(x=2.5, y=9.4, text="政策維持型", showarrow=False)
            matrix.add_annotation(x=7.5, y=0.8, text="効率改善優先", showarrow=False)
            matrix.add_annotation(x=2.5, y=0.8, text="再検討", showarrow=False)
            matrix.update_layout(title="経営判断マトリクス", xaxis_title="財務持続性", yaxis_title="政策・医療必要性", xaxis={"range": [0, 10]}, yaxis={"range": [0, 10]}, height=280, margin=dict(l=20, r=20, t=45, b=20))
            st.plotly_chart(matrix, use_container_width=True)

        rcol, ycol = st.columns(2)
        with rcol:
            breakdown = make_score_breakdown(result)
            radar = go.Figure()
            radar.add_trace(go.Scatterpolar(r=breakdown["スコア"].tolist() + [breakdown["スコア"].iloc[0]], theta=breakdown["指標"].tolist() + [breakdown["指標"].iloc[0]], fill="toself", name="評価"))
            radar.update_layout(title="評価内訳レーダー", polar={"radialaxis": {"visible": True, "range": [0, 10]}}, height=360, margin=dict(l=25, r=25, t=50, b=20))
            st.plotly_chart(radar, use_container_width=True)
        with ycol:
            base_margin = float(result.get("増員後粗利益人件費余力", 0) or 0)
            annual_cost = float(result.get("年間人件費増", 0) or 0)
            years = ["現在", "1年後", "2年後", "3年後"]
            margins = []
            for i in range(4):
                margins.append(base_margin - annual_cost * ((1 + inflation_rate) ** i - 1) * 0.35)
            trend = pd.DataFrame({"年度": years, "粗利益人件費余力": margins})
            fig = px.bar(trend, x="年度", y="粗利益人件費余力", title="3年粗利益余力シミュレーション") if px is not None else None
            if fig is not None:
                fig.update_layout(height=360, margin=dict(l=25, r=25, t=50, b=20))
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(trend.set_index("年度"))

        with st.expander("職種別ピア比較グラフ", expanded=True):
            fig = px.bar(comp, x="区分", y="対病床人員", title=f"{job} 対病床人員の比較") if px is not None else None
            if fig is not None:
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(comp.set_index("区分"))
    else:
        st.bar_chart(comp.set_index("区分"))


def render_side_console(
    target: pd.Series,
    peers: pd.DataFrame,
    peer_criteria: List[str],
    peer_applied: List[str],
    min_peers: int,
    result: Dict[str, object],
    hire_type_selected: str,
    request_type_selected: str,
    revenue_impact_selected: str,
    labor_mode: str,
    labor_shortage_factor: float,
    labor_weight: float,
    inflation_rate: float,
    inflation_cost_share: float,
) -> None:
    st.subheader("条件コンソール")
    with st.container(border=True):
        st.markdown("**ピア設定**")
        st.write(f"ピア件数: **{len(peers)}**")
        st.write(f"最小病院数: **{min_peers}**")
        st.write("抽出候補: " + (", ".join(peer_criteria) if peer_criteria else "未選択"))
        if peer_applied:
            st.write("適用条件:")
            for p in peer_applied:
                st.caption(f"・{p}")
        else:
            st.caption("件数不足等により全体比較または条件未適用です。")
        for col in ["過疎地区分", "小児フラグ", "精神科フラグ", "クラスタ", "DPCフラグ"]:
            if col in target.index:
                st.caption(f"{col}: {target.get(col)}")

    with st.expander("外部環境・前提条件", expanded=True):
        st.write({
            "労働市場の扱い": labor_mode,
            "人手不足逼迫度": round(float(labor_shortage_factor), 2),
            "人手不足リスク荷重": round(float(labor_weight), 2),
            "物価・賃金上昇圧力": f"{inflation_rate:.1%}",
            "物価上昇の人件費反映割合": f"{inflation_cost_share:.0%}",
        })

    with st.expander("主要算定結果", expanded=False):
        st.write({
            "年間人件費増": fmt_yen(result.get("年間人件費増")),
            "年間増収推計": fmt_yen(result.get("年間増収推計")),
            "政策・質安全価値": fmt_yen(result.get("政策・質安全価値")),
            "労働市場逼迫価値": fmt_yen(result.get("労働市場逼迫価値")),
            "粗利益持続性判定": result.get("粗利益持続性判定"),
        })

    st.write("")
    st.write("")
    with st.expander("申請フォーム反映内容（参考）", expanded=False):
        st.write({
            "採用区分": hire_type_selected or "未選択",
            "申請区分": request_type_selected or "未選択",
            "収支影響": revenue_impact_selected or "未選択",
            "年間増収推計_補正前": fmt_yen(result.get("年間増収推計_申請補正前")),
            "増収補正倍率": result.get("申請補正_増収倍率"),
            "政策・質安全価値_実効重み": result.get("政策・質安全価値_実効重み"),
            "労働市場逼迫度_実効値": result.get("労働市場逼迫度_実効値"),
        })
        st.markdown(str(result.get("申請区分補正メモ", "")).replace("\n", "  \n"))

def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title("職員人員要望意思決定支援シミュレーター")
    st.caption("経営・人員パネル、申請内容フォーム、厚生労働省 病床機能報告オープンデータを統合し、増員要望を病院別・職種別に評価します。")

    with st.sidebar:
        st.header("1. データ読込")
        financial_upload = st.file_uploader("経営・人員パネルxlsx（病院別の経営・職員数データ）", type=["xlsx"], key="financial")
        st.caption("申請表xlsxはアップロード不要です。病院別の経営・人員実績データだけを読み込みます。GitHub/Streamlit Cloudには個人情報・機微データを同梱しない設計です。")
        st.divider()
        st.header("2. シミュレーション設定")
        revenue_sensitivity = st.slider("不足解消による年間増収感応度", 0.0, 0.10, 0.015, 0.005, help="ピア中央値までの不足を埋めた場合、医業収益が何%改善するかの仮説")
        quality_weight = st.slider("政策・質安全価値の重み", 0.0, 1.0, 0.35, 0.05, help="収益化しにくい安全・政策必要性を人件費に対する便益として別枠換算")
        st.divider()
        st.header("3. 外部環境の重み")
        labor_mode = st.radio("労働市場の人手不足をどう扱うか", ["手動スライダー", "労働力トレンドを参考表示", "判定に入れない"], index=0)
        labor_shortage_manual = st.slider("人手不足の逼迫度", 0.0, 1.0, 0.55, 0.05, help="1に近いほど採用難・欠員長期化リスクを強く評価")
        labor_weight = st.slider("人手不足リスクの判定荷重", 0.0, 0.30, 0.15, 0.01)
        inflation_rate = st.slider("物価上昇率・賃金上昇圧力", -0.02, 0.10, 0.026, 0.001, format="%.3f", help="人件費や委託費等の将来コスト上昇を反映。例：0.026 = 2.6%")
        inflation_cost_share = st.slider("物価上昇を人件費増に反映する割合", 0.0, 1.0, 0.60, 0.05)
        st.divider()
        st.header("4. 表示設定")
        show_raw_table = st.checkbox("評価メモの詳細テーブルを表示", value=False)
        show_method_hint = st.checkbox("ダッシュボード内に算定式メモを表示", value=False)

    financial_bytes = require_excel_bytes(read_file_bytes(financial_upload, DEFAULT_FINANCIAL_FILE), "経営・人員パネルxlsx")
    raw = load_financial_excel(financial_bytes)
    df = add_core_metrics(raw)
    df_latest = add_core_metrics(latest_rows(raw))

    tab_application, tab_decision, tab_data, tab_mhlw, tab_method = st.tabs([
        "申請入力・確認", "意思決定ダッシュボード", "経営・人員データ", "病床機能報告マージ", "算定ロジック",
    ])

    with tab_application:
        app = application_input_panel()
        st.divider()
        st.subheader("申請内容の確認")
        app_table = pd.DataFrame([{"項目": k, "内容": v} for k, v in app.items() if not isinstance(v, dict)])
        st.dataframe(app_table, hide_index=True, use_container_width=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**採用区分**")
            st.json(app.get("採用区分_候補", {}))
        with c2:
            st.markdown("**申請区分**")
            st.json(app.get("申請区分_候補", {}))
        with c3:
            st.markdown("**収支影響**")
            st.json(app.get("収支影響_候補", {}))

    with tab_decision:
        if "病院名称" not in df_latest.columns:
            st.error("経営・人員パネルxlsxに '病院名称' 列が必要です。")
            st.stop()

        main_col = st.container()
        with main_col:
            with st.expander("対象・申請人数の設定", expanded=True):
                h1, h2, h3 = st.columns([1.4, 1.0, 1.0])
                names = sorted(df_latest["病院名称"].dropna().astype(str).unique())
                default_hospital = fuzzy_pick_hospital(df_latest, str(app.get("所属名", "")))
                default_index = names.index(default_hospital) if default_hospital in names else 0
                with h1:
                    hospital = st.selectbox("対象病院", names, index=default_index)
                target = df_latest[df_latest["病院名称"] == hospital].iloc[0]
                raw_job = str(app.get("希望職種_raw") or "")
                inferred_job = classify_job(raw_job)
                with h2:
                    job = st.selectbox("評価する職種", list(STAFF_COL_BY_JOB.keys()), index=list(STAFF_COL_BY_JOB.keys()).index(inferred_job))
                proposed_count = app.get("希望人数") or 1.0
                with h3:
                    add_count = st.number_input("申請人数 / シミュレーション増員数", 0.0, 100.0, float(proposed_count), 0.5)
                annual_cost = st.number_input("1人あたり年間人件費（円）", 0, 50_000_000, DEFAULT_ANNUAL_COST_YEN[job], 100_000)

            with st.expander("ピア抽出条件を変更", expanded=False):
                available_peer_cols = [c for c in ["クラスタ", "DPCフラグ", "過疎地区分", "小児フラグ", "精神科フラグ", "都道府県名", "二次医療圏名"] if c in df_latest.columns]
                default_peer_cols = [c for c in ["クラスタ", "過疎地区分", "小児フラグ", "精神科フラグ"] if c in available_peer_cols]
                peer_criteria = st.multiselect("ピア抽出に使う属性", available_peer_cols, default=default_peer_cols, help="過疎地区分・小児フラグ・精神科フラグを含めて、対象病院と同じ属性の病院群に絞り込みます。")
                min_peers = st.slider("ピア条件を採用する最小病院数", 3, 30, 5, 1)

        # expander内で定義される値を安全にするための初期値
        if 'peer_criteria' not in locals():
            available_peer_cols = [c for c in ["クラスタ", "DPCフラグ", "過疎地区分", "小児フラグ", "精神科フラグ", "都道府県名", "二次医療圏名"] if c in df_latest.columns]
            peer_criteria = [c for c in ["クラスタ", "過疎地区分", "小児フラグ", "精神科フラグ"] if c in available_peer_cols]
            min_peers = 5
        peers, peer_applied = peer_group(df_latest, target, peer_criteria, min_peers=min_peers)

        hire_type_selected = selected_key(app.get("採用区分_候補", {}))
        request_type_selected = selected_key(app.get("申請区分_候補", {}))
        revenue_impact_selected = selected_key(app.get("収支影響_候補", {}))
        labor_shortage_factor = 0.0 if labor_mode == "判定に入れない" else labor_shortage_manual

        result = simulate(
            target, peers, job, add_count, annual_cost, revenue_sensitivity, quality_weight,
            labor_shortage_factor=labor_shortage_factor,
            labor_weight=labor_weight,
            inflation_rate=inflation_rate,
            inflation_cost_share=inflation_cost_share,
            hire_type=hire_type_selected,
            request_type=request_type_selected,
            revenue_impact=revenue_impact_selected,
        )

        comp = pd.DataFrame({
            "区分": ["対象病院 現在", "申請後", "ピア中央値"],
            "対病床人員": [result["現員_対病床"], result["増員後_対病床"], result["ピア中央値_対病床"]],
        })

        with main_col:
            render_visual_dashboard(result, job, hospital, comp, labor_shortage_factor, inflation_rate)

            with st.expander("粗利益・人件費の詳細", expanded=False):
                g1, g2, g3, g4 = st.columns(4)
                g1.metric("現在粗利益", fmt_yen(result['現在粗利益']))
                g2.metric("増員後人件費", fmt_yen(result['増員後人件費']))
                cover = result['増員後粗利益人件費カバー率']
                g3.metric("増員後 粗利益/人件費", "-" if pd.isna(cover) else f"{cover:.2f}倍")
                g4.metric("年間人件費増", fmt_yen(result['年間人件費増']))
                st.info(result['粗利益持続性判定'])

            if show_raw_table:
                with st.expander("評価メモ 詳細テーブル", expanded=True):
                    memo = pd.DataFrame([result]).T.rename(columns={0: "値"})
                    st.dataframe(memo, use_container_width=True)

            if show_method_hint:
                with st.expander("算定式メモ", expanded=True):
                    st.markdown("""
                    - 粗利益 = 修正医業収益 -（医薬品費 + 医療材料費）
                    - 財務効果 = 年間増収推計 - 物価反映後の年間人件費増
                    - 政策・労働市場価値込み純効果 = 年間増収推計 + 政策・質安全価値 + 労働市場逼迫価値 - 年間人件費増
                    - 最終スコア = 人員不足解消、財務効果、粗利益持続性、政策必要性、労働市場逼迫度の加重平均
                    """)

        with st.sidebar:
            st.divider()
            render_side_console(
                target, peers, peer_criteria, peer_applied, min_peers, result,
                hire_type_selected, request_type_selected, revenue_impact_selected,
                labor_mode, labor_shortage_factor, labor_weight, inflation_rate, inflation_cost_share,
            )

    with tab_data:
        st.subheader("経営・人員データ")
        prefectures = sorted(df["都道府県名"].dropna().astype(str).unique()) if "都道府県名" in df.columns else []
        selected_pref = st.multiselect("都道府県フィルタ", prefectures, default=["長野県"] if "長野県" in prefectures else prefectures[:1])
        view = df[df["都道府県名"].isin(selected_pref)] if selected_pref and "都道府県名" in df.columns else df
        show_cols = [c for c in ["年度", "法人名称等", "病院名称", "都道府県名", "病床数", "全職員数", "医師数", "看護師数", "医療スタッフ数", "事務職員数", "修正医業収益", "医業費用", "人件費", "医薬品費", "医療材料費", "粗利益_円", "粗利益人件費カバー率", "病床稼働率", "職員数_対病床", "医師_対病床", "看護師_対病床", "医療スタッフ_対病床", "過疎地区分", "小児フラグ", "精神科フラグ", "クラスタ"] if c in view.columns]
        sort_cols = [c for c in ["病院名称", "年度"] if c in show_cols]
        st.dataframe(view[show_cols].sort_values(sort_cols) if sort_cols else view[show_cols], use_container_width=True, height=520)
        csv = view[show_cols].to_csv(index=False).encode("utf-8-sig")
        st.download_button("表示データをCSVで出力", data=csv, file_name="filtered_staff_simulation_data.csv", mime="text/csv")

    with tab_mhlw:
        st.subheader("厚生労働省 病床機能報告オープンデータ")
        st.write("令和6年度病床機能報告ページからExcelリンクを取得し、必要に応じて病棟票データを読み込めます。病院名・医療機関番号で突合する前提です。")
        links = scrape_mhlw_links()
        st.dataframe(links, hide_index=True, use_container_width=True)
        st.caption("長野県は中部地方ファイルに含まれるため、既定で中部地方の病棟票を参照します。環境によりMHLWサイトへの接続が失敗する場合は、上表のURLからxlsxを取得して手動アップロードしてください。")
        mhlw_url = st.text_input("読み込むMHLW xlsx URL", MHLW_R6_CHUBU_WARD_STYLE1)
        if st.button("MHLWデータを試読込"):
            mhlw_df = try_load_mhlw_excel(mhlw_url)
            st.dataframe(mhlw_df.head(100), use_container_width=True)
            st.info("実運用では、病院名称または医療機関番号の列名を確認し、経営パネルと左結合してください。列名が年度により変わるため、この画面で列を確認できるようにしています。")

    with tab_method:
        st.subheader("算定・判定ロジック")
        st.markdown(
            """
            1. 経営・人員パネルから、対象病院の最新年度を抽出します。  
            2. クラスタ、DPC、過疎地区分、小児フラグ、精神科フラグなど、UIで選んだ属性からピア群を作り、職種別の「対病床人員」中央値を算定します。  
            3. 申請人数により、対象病院の対病床人員がピア中央値にどれだけ近づくかを評価します。  
            4. 粗利益は `修正医業収益 -（医薬品費 + 医療材料費）` とし、増員後に粗利益で人件費を支えられるかを `粗利益 / 人件費` と余力額で判定します。  
            5. 財務効果は `医業収益 × 増収感応度 × 不足解消率 - 物価反映後の年間人件費増` で試算します。  
            6. 政策・質安全価値は、収益化しにくい必要性を `年間人件費増 × 重み` として別枠表示します。  
            7. 労働市場の人手不足は、手動スライダーまたは労働力トレンドを参考にした荷重として、採用難・欠員長期化の機会損失に反映します。  
            8. 物価上昇率・賃金上昇圧力は、将来コスト上昇として人件費増に反映します。  
            9. 最終スコアは、人員不足解消・財務効果・粗利益持続性・政策必要性・労働市場逼迫度の加重平均です。係数は意思決定会議で調整してください。
            """
        )
        st.subheader("外部統計の根拠URL")
        st.markdown(f"- 労働力人口・労働力トレンド: {JILPT_LABOR_FORCE_URL}")
        st.markdown(f"- 消費者物価指数（CPI）: {STAT_CPI_URL}")
        st.markdown(f"- CPI最新月次結果: {STAT_CPI_LATEST_URL}")
        st.warning("このアプリは意思決定を支援する試算モデルです。最終判断では、施設基準、夜勤配置、地域医療構想、採用市場、予算制約を併せて確認してください。")


if __name__ == "__main__":
    main()
